import asyncio
from asyncio.windows_events import NULL
import random
from typing import List
import discord
from discord.ext import commands
import openpyxl
import os
import openpyxl as OP
import openpyxl.worksheet
import xlwings

import datetime
import time
from Token import Token

Intents_=discord.Intents.default()
Intents_.message_content = True
bot = commands.Bot(command_prefix='$', help_command=None, intents=Intents_)
excelfilename = 'HabitTrackerDiscordBot.xlsx'

wb = OP.load_workbook(os.path.join(os.path.dirname(__file__),excelfilename), data_only=True)
ws = wb['BaseRef']

print("Up and running")

PreviousDay = None

def dailygoalcheck():
    global PreviousDay

    if(PreviousDay == None):
        PreviousDay = ws['B11'].value

    CurrentDay = datetime.datetime.today().day

    if(CurrentDay > PreviousDay):
        PreviousDay = CurrentDay
        if(ws['B9'].value > 0):
            ws['B11'].value = CurrentDay
            for i in range(1, ws['B9'].value):
                username : str = ws['H' + str(i + 1)].value
                usersheet = wb[username]
                usersheet['C3'].value = 0
            wb.save(excelfilename)
    
    time.sleep(60*60)

    dailygoalcheck()
    
@bot.command()
async def help(ctx):
    embed = discord.Embed(title="All commands", colour=discord.Colour(0x3e038c))
    embed.add_field(name=f"Command", value="".join(map(str,outputtable(9, 2, 1, 11, ws))), inline=True)
    embed.add_field(name=f"Description", value="".join(map(str,outputtable(10, 2, 1, 11, ws))), inline=True)
    await ctx.send(embed=embed)

def numtochar(input:int):
    start_index = 1   #  it can start either at 0 or at 1
    letter = ''
    while input > 25 + start_index:
        letter += chr(65 + int((input-start_index)/26) - 1)
        input = input - (int((input-start_index)/26))*26
    letter += chr(65 - start_index + (int(input)))
    return letter

def outputtable(colstart:int, rowstart:int, colsize:int, rowsize:int, sheet):
    Output = []
    for row in range(rowstart, rowstart + rowsize):
        for col in range(colstart, colstart + colsize):
            Output.append(sheet[numtochar(col) + str(row)].value)
            if((col - colstart) < colsize - 1):
                Output.append(" : ")
        Output.append("\n")
    return Output

def finduser(name:str):
    if(ws['B9'].value > 0):
        for i in range(1, ws['B9'].value):
                if(ws['H' + str(i + 1)].value == name):
                    return i
    return 0

## USER COMMANDS

@bot.command()
@commands.cooldown(1, 5, commands.BucketType.user)
async def adduser(ctx):
    ctx.typing()
    
    username = ctx.author.name
    result = finduser(username)
    if(result != 0):
        await ctx.send(username + " already exists as a current user")
    else:
        ws.move_range("H2:H"+str(ws['B9'].value + 1), rows = 1, cols = 0)
        ws['B9'].value += 1
        wb.save(excelfilename)

        ws['H2'].value = username

        newsheet = wb.copy_worksheet(wb['Username'])
        newsheet.title = username

        wb.save(excelfilename)
        
        await ctx.send("User " + username + " added to the list")

@bot.command()
@commands.cooldown(1, 5, commands.BucketType.user)
async def removeuser(ctx):
    ctx.typing()
    result = finduser(ctx.author.name)
    if(result == 0):
        ctx.send("User not found")
    else:
        await ctx.send("Do you really want to remove yourself from current users? (yes/no)(y/n)") 

        def check(m):
            return m.author == ctx.author and m.channel == ctx.channel
    
        try:
            response = await bot.wait_for('message', check=check, timeout = 30.0)
        except asyncio.TImeoutError:
            await ctx.send("Time's up :/")
            return 
        
        if response.content.lower() not in ("yes", "y"):
            await ctx.send("Cancelled")
            return
        
        del wb[ctx.author.name]
        ws.move_range("H" + str(result + 2) + ":H" + str(ws['B9'].value + result + 2), rows=-1, cols=0)
        ws['B9'].value -= 1
        wb.save(excelfilename)
        await ctx.send("User successfully removed")

@bot.command()
async def add(ctx, duration:int = -1, importance:int = -1,repeat:int = 0, desc:str = ""):
    global wb
    global ws

    ctx.typing()

    if(duration == -1) or (importance == -1):
        await ctx.send("Error: Duration and Importance are mandatory values")
    else:
        await ctx.send("Enter the task name") 

        def check(m):
            return m.author == ctx.author and m.channel == ctx.channel
    
        try:
            response = await bot.wait_for('message', check=check, timeout = 30.0)
        except asyncio.TImeoutError:
            await ctx.send("You took too long to respond!")
            return 

        username = ctx.author.name
        usersheet = wb[username]

        usersheet.move_range("D2:J"+str(ws['B5'].value + 1), rows = 1, cols = 0, translate=True)
        usersheet['B5'].value += 1
        wb.save(excelfilename)

        usersheet['D2'].value = response.content
        usersheet['E2'].value = duration
        usersheet['G2'].value = importance
        usersheet['H2'].value = repeat

        usersheet['J2'] = "=BaseRef!$B$1+(" + username + "!E2 * BaseRef!$B$3)+(" + username + "!G2 * BaseRef!$B$5)"

        wb.save(excelfilename)

        excelapp = xlwings.App(visible = False)
        excelbook = excelapp.books.open(excelfilename)
        excelbook.save()
        excelbook.close()
        excelapp.quit()

        wb = OP.load_workbook(os.path.join(os.path.dirname(__file__),excelfilename), data_only=True)
        ws = wb['BaseRef']
        
        await ctx.send("Task added to the list")

@bot.command()
async def remove(ctx):
    ctx.typing()

    await ctx.send("Enter the task name") 

    def check(m):
        return m.author == ctx.author and m.channel == ctx.channel

    try:
        response = await bot.wait_for('message', check=check, timeout = 30.0)
    except asyncio.TImeoutError:
        await ctx.send("You took too long to respond!")
        return 

    usersheet = wb[ctx.author.name]
    result = -1

    for row in range(1, usersheet['B5'].value):
        if(response.content.lower() == usersheet['D' + str(row + 1)].value.lower()):
            result = usersheet['J' + str(row + 1)].value
            usersheet.move_range("D" + str(row + 2) + ":J" + str(usersheet['B5'].value + row + 2), rows=-1, cols=0)
            usersheet['B5'].value -= 1
            wb.save(excelfilename)

    if(result != -1):
        await ctx.send("Task successfully removed")
    else:
        await ctx.send("Task not found")

@bot.command()
async def setdailygoal(ctx, amnt:int):
    ctx.typing()

    if(finduser(ctx.author.name) == 0):
        await ctx.send("**Error** : User not found")
    else:
        wb[ctx.author.name]['B3'].value = amnt
        await ctx.send("Daily goal modified")

@bot.command()
async def finish(ctx):
    ctx.typing()

    await ctx.send("Enter the task name") 

    def check(m):
        return m.author == ctx.author and m.channel == ctx.channel

    try:
        response = await bot.wait_for('message', check=check, timeout = 30.0)
    except asyncio.TImeoutError:
        await ctx.send("You took too long to respond!")
        return 
    
    usersheet = wb[ctx.author.name]
    result = -1

    for row in range(1, usersheet['B5'].value):
        if(response.content.lower() == usersheet['D' + str(row + 1)].value.lower()):
            if(usersheet['H' + str(row + 1)].value > 0):
                usersheet['H' + str(row + 1)].value -= 1
            elif(usersheet['H' + str(row + 1)].value != -1):
                usersheet.move_range("D" + str(row + 2) + ":J" + str(usersheet['B5'].value + row + 2), rows=-1, cols=0)
                usersheet['B5'].value -= 1
            wb.save(excelfilename)
            result = usersheet['J' + str(row + 1)].value

    if(result != -1):
        usersheet = wb[ctx.author.name]
        usersheet['B1'].value += result
        usersheet['B2'].value += 1 
        usersheet['C3'].value += 1

        if(usersheet['C3'].value >= usersheet['B3'].value):
            await ctx.send("Daily goal reached! You got a bonus :D")
            usersheet['B1'].value += ws['B1'].value * ws['B1'].value

        wb.save(excelfilename)
        # add to daily goal

        await ctx.send("Task completed!")
    else:
        await ctx.send("Task not found")

## DISPLAY COMMANDS

@bot.command()
async def tasks(ctx):
    ctx.typing()

    targetuser: str
    if(ctx.message.mentions):
        targetuser = ctx.message.mentions[0].name
    else:
        targetuser = ctx.author.name
    
    if(finduser(targetuser) == 0):
        await ctx.send("**Error** : User not found")
    else:
        usersheet = wb[targetuser]
        amnt = usersheet["B5"].value

        embed = discord.Embed(title=targetuser + "'s tasks", colour=discord.Colour(0x3e038c),)
        embed.add_field(name=f"Name", value="".join(map(str,outputtable(4, 2, 1, amnt, usersheet))), inline=True)
        embed.add_field(name=f"Importance", value="".join(map(str,outputtable(7, 2, 1, amnt, usersheet))), inline=True)
        embed.add_field(name=f"Repeat", value="".join(map(str,outputtable(8, 2, 1, amnt, usersheet))), inline=True)
        embed.add_field(name=f"Points", value="".join(map(str,outputtable(10, 2, 1, amnt, usersheet))), inline=True)
        await ctx.send(embed=embed)

@bot.command()
async def shop(ctx):
    embed = discord.Embed(title="Point shop", description="use $buy [ItemName] to purchase an item from below", colour=discord.Colour(0x3e038c))
    embed.add_field(name=f"Name", value="".join(map(str,outputtable(4, 2, 1, 5, ws))), inline=True)
    embed.add_field(name=f"Price", value="".join(map(str,outputtable(5, 2, 1, 5, ws))), inline=True)
    embed.add_field(name=f"Description", value="".join(map(str,outputtable(6, 2, 1, 5, ws))), inline=True)
    await ctx.send(embed=embed)

@bot.command()
async def buy(ctx, item:str, amnt:int = 1):
    bought = False
    if(amnt < 1):
        await ctx.send("Item amount can't be negative >:[")
        return
    
    for row in range(1, ws['B8'].value):
        if(ws["D"+ str(row + 1)].value.lower() == item.lower()):
            usersheet = wb[ctx.author.name]
            price = ws["E" + str(row + 1)].value

            if(usersheet["B1"].value >= price * amnt):
                usersheet["M" + str(row + 1)].value += amnt
                usersheet["B1"].value -= price * amnt
                wb.save(excelfilename)
                bought = True
            else:
                await ctx.send("You don't have enough points :(")
                return
    
    if(bought):
        await ctx.send("Successfully bought the item")
    else:
        await ctx.send("Item not found")

@bot.command()
async def stats(ctx):
    ctx.typing()

    targetuser: str
    if(ctx.message.mentions):
        targetuser = ctx.message.mentions[0].name
    else:
        targetuser = ctx.author.name
    
    if(finduser(targetuser) == 0):
        await ctx.send("**Error** : User not found")
    else:
        embed = discord.Embed(title=targetuser + "'s stats", colour=discord.Colour(0x3e038c))
        embed.add_field(name=f"", value="".join(map(str,outputtable(1, 1, 1, 5, wb[targetuser]))), inline=True)
        embed.add_field(name=f"", value="".join(map(str,outputtable(2, 1, 1, 5, wb[targetuser]))), inline=True)
        await ctx.send(embed=embed)

        embed_ = discord.Embed(title=targetuser + "'s items", colour=discord.Colour(0x3e038c))
        embed_.add_field(name=f"", value="".join(map(str,outputtable(12, 2, 1, 5, wb[targetuser]))), inline=True)
        embed_.add_field(name=f"", value="".join(map(str,outputtable(13, 2, 1, 5, wb[targetuser]))), inline=True)
        await ctx.send(embed=embed_)

## Admin commands
@bot.command()
@commands.has_permissions(administrator=True)
async def pointsettings(ctx):
    ctx.typing()
    embed = discord.Embed(title="Point settings", colour=discord.Colour(0x3e038c))

    embed.add_field(name=f"", value="".join(map(str,outputtable(1, 1, 2, 5, ws))), inline=False)

    await ctx.send(embed=embed)

bot.run(Token)
# # Don't reveal your bot token, regenerate it asap if you do

dailygoalcheck()